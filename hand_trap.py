from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from pathlib import Path

def insert_image(card_image, location):
    image_extensions = [".png", ".jpg", ".jpeg"]
    for ext in image_extensions:
        image_file_path = image_folder / f"{card_image}{ext}"
        # 이미지 파일이 존재하면 엑셀에 추가
        if image_file_path.exists():
            # 이미지 객체 생성
            img = Image(str(image_file_path))

            # 이미지 크기 조정
            cell_width = 15
            cell_height = 130
            img.width = cell_width * 8
            img.height = cell_height * 1.33
            ws.add_image(img, location)
            return True
    print(f"이미지를 찾을 수 없습니다: {card_image}")
    return False

def trap_input(input_text, input_type):
    """
    input_type : 0이면 split(), 1이면 그대로 
    """
    user_input = input(input_text)
    my_trap.append(user_input)
    if input_type:
        return user_input
    else:
        return user_input.split()
    
file_name = trap_input("수정할 파일명을 입력하세요 (확장자 제외): ", 1) + ".xlsx"
image_folder = Path.cwd() / "이미지" / file_name.split(".xlsx")[0]
my_trap = list()

try:
    # 기존 파일 열기
    wb = load_workbook(file_name)
    print(f"'{file_name}' 파일을 불러왔습니다.")
except FileNotFoundError:
    # 파일이 없을 경우 새 파일 생성
    wb = Workbook()
    print(f"'{file_name}' 파일이 없어 새 파일을 생성합니다.")

# 기존 시트 선택 또는 새로운 시트 생성
sheet_name = file_name + "패트랩 타이밍"
my_trap.append(sheet_name)
if sheet_name in wb.sheetnames:
    # 기존 시트 삭제
    wb.remove(wb[sheet_name])
    print(f"'{sheet_name}' 시트를 삭제하고 새로 생성합니다.")
ws = wb.create_sheet(sheet_name)