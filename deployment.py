# 함수로 모듈화하기
# 함수에 docstring 작성
# 불필요한 코드 제거
# 전개 텍스트 읽어서 만드는 기능
# 전개 결과물 수비 표시 표현

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from pathlib import Path
from itertools import count


def insert_image(card_image, location):
    """
    card_image: 이미지로 넣을 대상
    
    location: 이미지가 들어갈 위치
    
    엑셀 파일과 이름이 같은 폴더에 접근하여 이미지가 존재한다면 location을 참고하여 그 위치에 card_image의
    이미지 파일을 엑셀에 삽입
    """
    image_extensions = [".png", ".jpg", ".jpeg"]
    for ext in image_extensions:
        image_file_path = image_folder / f"{card_image}{ext}"
        # 이미지 파일이 존재하면 엑셀에 추가
        if image_file_path.exists():
            # 이미지 객체 생성
            img = Image(str(image_file_path))

            # 이미지 크기 조정
            cell_width = 15
            cell_height = 130
            img.width = cell_width * 8
            img.height = cell_height * 1.33
            ws.add_image(img, location)
            return True
    print(f"이미지를 찾을 수 없습니다: {card_image}")
    return False

def name_and_way():
    """
    입력 받은 내용을 엑셀에 옮기기 쉽도록 가공하는 함수
    """
    big_data = [[], []]
    while True:
        # 0번에 카드명, 1번에 방법
        name_way_data = deployment_input("카드명과 방법을 입력하세요 (종료하려면 빈 입력): ", 0, my_deployment)
        if not name_way_data:
            try:
                big_data[0].pop()
                big_data[1].pop()
            except IndexError:
                pass
            break
        if name_way_data[0] == 'and' or name_way_data[0] == 'or':
            big_data[0].pop()
            big_data[0].append(name_way_data[0])
        else:
            for name_way_idx, name_way_value in enumerate(name_way_data):
                if name_way_idx % 2 == 0:
                    big_data[0].append(name_way_value.replace("_", " "))
                else:
                    big_data[1].append(name_way_value.replace("_", " "))
            big_data[0].append(">")
            big_data[1].append(" ")
    return big_data

def insert_deployment(ID_row_idx, ID_col_idx):
    """
    ID_row_idx: 기준 row 위치
    
    ID_col_idx: 기준 col 위치
    
    기준 위치에서부터 입력 받은 내용을 두줄로 엑셀로 옮기고 다음 입력이 기준 값(현재 15)을 넘으면 줄바꿈 처리
    """
    start_ID_row_idx = ID_row_idx
    start_ID_col_idx = ID_col_idx
    big_data = name_and_way()

    for ID_i, ID_data in enumerate(big_data):
        for ID_idx ,ID_value in enumerate(ID_data):
            if ((ID_value == ">" or ID_value == ' ' or ID_value == "or" or ID_value == "and") and find_location(ID_idx, ID_col_idx, [">", " ", "or", "and"], ID_data, 15)) or ID_col_idx > 15:
                ID_col_idx = 1
                ID_row_idx += 2

            ID_cell_location = f"{chr(65 + ID_col_idx - 1)}{ID_row_idx}"
            if not insert_image(ID_value, ID_cell_location):
                ws.cell(row=ID_row_idx, column=ID_col_idx, value=ID_value)
            ID_col_idx += 1
        # 첫번째 for문이 끝나고 방법에 해당하는 반복문이 진행되기 전에 row, col의 정보를 갱신
        if ID_i == 0:
            ID_row_idx, ID_col_idx = start_ID_row_idx+1, start_ID_col_idx
            
    for ID_row in range(start_ID_row_idx, ID_row_idx, 2):
        ws.row_dimensions[ID_row].height = 130
    for ID_enu, ID_row in enumerate(ws.iter_rows(min_row=start_ID_row_idx, max_row=ID_row_idx, min_col=1, max_col=ws.max_column), start=1):
        for ID_cell in ID_row:
            ID_cell.alignment = alignment_style
            if ID_enu % 2 == 0:
                ID_cell.font = font_style_text
            else:
                ID_cell.font = font_style_card

def find_location(current_target_idx=int, current_col_idx=int, targets=list, target_list=list, search_range=int):
    '''
    current_target_idx: 앞선 column 중 가장 가까운 target 위치
    
    current_col_idx: 현재 column의 위치
    
    targets: 찾을 내용
    
    target_list: target이 담겨있는 list의 내용
    
    search_range: target이 몇 column뒤 까지 있는지 찾을 범위
    
    target_list에서 다음 targets 중 가장 가까운 target이 search_range 이내에 존재하는지 파악 후
    다음 target에서 column위치가 search_range를 넘으면 True, 아니면 False 반환
    '''
    next_target_idx_list = list()
    next_target_idx_list.append(len(target_list)-1)
    for target in targets:
        try:
            next_target_idx_list.append(target_list.index(target, current_target_idx + 1) - 1)
        except ValueError:
            pass
    if min(next_target_idx_list) - current_target_idx + current_col_idx > search_range:
        return True
    return False

# 전개 텍스트 저장 함수
def deployment_save(sheet, deployment):
    """
    sheet: 텍스트가 저장될 sheet의 이름
    
    deployment: sheet에 저장될 텍스트
    """
    wb = load_workbook(file_path)
    ws = wb[sheet]
    for i in count(start=1):
        check = ws.cell(row=1, column=i).value
        if check is None or check == deployment[1]:
            ws.cell(row=1, column=i, value=sheet_name)
            ws.cell(row=2, column=i, value="\n".join(map(str, deployment)))
            wb.save(file_path)
            return print(f"{chr(64 + i)}번에 {sheet_name} 전개 텍스트 저장")

def deployment_input(input_text, input_type, save_list):
    """
    input_text: input을 사용할 때 이용자에게 보여줄 메시지
     
    input_type: 0이면 split(), 1이면 그대로 
    
    save_list: input으로 입력받은 내용이 저장될 list
    
    deployment_save에서 사용하기 위해 텍스트를 save_list에 저장하고 반환하는 함수
    """
    user_input = input(input_text)
    save_list.append(user_input)
    if input_type:
        return user_input.replace("_", " ")
    else:
        return [word.replace("_", " ") for word in user_input.split()]

my_deployment = list()
current_dir = Path.cwd()
file_name = deployment_input("수정할 파일명을 입력하세요 (확장자 제외): ", 1, my_deployment)
image_folder = current_dir / "이미지" / file_name.split(".xlsx")[0]
deployment_folder = current_dir / "전개법" / file_name.split(".xlsx")[0]
file_path = deployment_folder.with_suffix(".xlsx")

# 폰트 설정
alignment_style = Alignment(horizontal="center", vertical="center")
font_style_card = Font(bold=True, size=30)
font_style_text = Font(bold=True, size=11)

extra_color = PatternFill(start_color="B3CEFB", end_color="B3CEFB", fill_type="solid")
monster_color = PatternFill(start_color="FFC599", end_color="FFC599", fill_type="solid")
magic_trap_color = PatternFill(start_color="A6E3B7", end_color="A6E3B7", fill_type="solid")
tomb_color = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

border_style = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

try:
    # 기존 파일 열기
    wb = load_workbook(file_path)
    print(f"'{file_path}' 파일을 불러왔습니다.")
except FileNotFoundError:
    # 파일이 없을 경우 새 파일 생성
    wb = Workbook()
    print(f"'{file_path}' 파일이 없어 새 파일을 생성합니다.")

# 기존 시트 선택 또는 새로운 시트 생성
sheet_name = deployment_input("작성할 시트명을 입력하세요: ", 1, my_deployment)
if sheet_name in wb.sheetnames:
    # 기존 시트 삭제
    wb.remove(wb[sheet_name])
    print(f"'{sheet_name}' 시트를 삭제하고 새로 생성합니다.")
ws = wb.create_sheet(sheet_name)
    
# 시트명을 A1에 입력
ws.cell(row=1, column=1, value=sheet_name)
ws.cell(row=2, column=1, value="핸드")
ws.cell(row=3, column=1, value="전개")
# 핸드 내용을 입력
for col_idx, value in enumerate(sheet_name.split(), start=2):
    cell_location = f"{chr(65 + col_idx - 1)}{2}"
    if not insert_image(value, cell_location):
        ws.cell(row=2, column=col_idx, value=value)

# 전개
insert_deployment(3, 2)

# 엔드 페이즈
q = deployment_input("엔드 페이즈 행동을 만드시겠습니까? (넘어가려면 빈 입력): " , 1, my_deployment)
if q:
    for i in count(start=1):
        if ws.cell(row=ws.max_row, column=i).value is None:
            ws.cell(row=ws.max_row+1, column=1, value="엔드")
            ws.cell(row=ws.max_row+1, column=1, value=" ")
            insert_deployment(ws.max_row-1, 2)
            break
# 결과 필드 만들기
ws.cell(row=ws.max_row+1, column=1, value="결과")
result_row = ws.max_row
# 엑스트라 몬스터
ws.cell(row=result_row, column=4).fill = extra_color
ws.cell(row=result_row, column=4).border = border_style
ws.cell(row=result_row, column=6).fill = extra_color
ws.cell(row=result_row, column=6).border = border_style

# 몬스터
for col in ws.iter_rows(min_row=result_row+1, max_row=result_row+1, min_col=3, max_col=7):
    for cell in col:
        cell.fill = monster_color
        cell.border = border_style
# 마함
for col in ws.iter_rows(min_row=result_row+2, max_row=result_row+2, min_col=3, max_col=7):
    for cell in col:
        cell.fill = magic_trap_color
        cell.border = border_style
# 필마
ws.cell(row=result_row+1, column=2).fill = magic_trap_color
ws.cell(row=result_row+1, column=2).border = border_style

# 묘지&제외
for row in ws.iter_rows(min_row=result_row, max_row=result_row+1, min_col=8, max_col=8):
    for cell in row:
        cell.fill = tomb_color
        cell.border = border_style

# 결과물 이미지 붙이기
q = deployment_input("결과 필드를 만드시겠습니까? (넘어가려면 빈 입력): ", 1, my_deployment)
if q:
    result_data = [[], []]
    for i in range(2):
        if i == 0:
            result = deployment_input("왼쪽 엑스트라 몬스터 존에 있는 카드를 입력해주세요 (넘어가려면 빈 입력): ", 1, my_deployment)
            if result:
                result_data[0].append(result)
                result_data[1].append([result_row, 4])
        else:
            result = deployment_input("오른쪽 엑스트라 몬스터 존에 있는 카드를 입력해주세요 (넘어가려면 빈 입력): ", 1, my_deployment)
            if result:
                result_data[0].append(result)
                result_data[1].append([result_row, 6])
                
    for i in range(1, 6):
        result = deployment_input(f"{i}번 몬스터 존에 있는 카드를 입력해주세요 (넘어가려면 빈 입력): ", 1, my_deployment)
        if result:
            result_data[0].append(result)
            result_data[1].append([result_row+1, 2+i])
            
    for i in range(1, 6):
        result = deployment_input(f"{i}번 마법 & 함정 존에 있는 카드를 입력해주세요 (넘어가려면 빈 입력): ", 1, my_deployment)
        if result:
            result_data[0].append(result)
            result_data[1].append([result_row+2, 2+i])
    
    result = deployment_input("필드 마법 존에 있는 카드를 입력해주세요 (넘어가려면 빈 입력): ", 1, my_deployment)
    if result:
        result_data[0].append(result)
        result_data[1].append([result_row+1, 2])
        
    result = deployment_input("엑스트라 덱에 있는 카드를 입력해주세요 (넘어가려면 빈 입력): ", 1, my_deployment)
    if result:
        result_data[0].append(result)
        result_data[1].append([result_row+2, 2])
    else:
        result_data[0].append("패")
        result_data[1].append([result_row+2, 2])
    
    for i in count():
        result = deployment_input("제외 존에 보여주고 싶은 카드를 입력해주세요 (넘어가려면 빈 입력): ", 1, my_deployment)
        if not result:
            break
        else:
            result_data[0].append(result)
            result_data[1].append([result_row, 8+i])
        
    for i in count():
        result = deployment_input("묘지 존에 보여주고 싶은 카드를 입력해주세요 (넘어가려면 빈 입력): ", 1, my_deployment)
        if not result:
            break
        else:
            result_data[0].append(result)
            result_data[1].append([result_row+1, 8+i])
            
    # 덱 표시
    result_data[0].append("패")
    result_data[1].append([result_row+2, 8])
    
    # 패 상태 입력
    result = deployment_input("남은 패 매수를 입력해주세요 (넘어가려면 빈 입력): ", 1, my_deployment)
    if result:
        hand_result = [[], []]
        hand_count = int(result)
        for i in range(1, hand_count+1):
            result = deployment_input("패에 특정 카드가 존재한다면 카드명과 방법을 입력해주세요 (넘어가려면 빈 입력): ", 0, my_deployment)
            if not result:
                for j in range(1, hand_count-(i-1)+1):
                    hand_result[0].append("패")
                    hand_result[1].append(" ")
                break
            hand_result[0].append(result[0].replace("_", " "))
            hand_result[1].append(result[1].replace("_", " "))
        
        for i, data in enumerate(hand_result):
            for j, value in enumerate(data, start=3):
                row_idx, col_idx = result_row+3, j
                cell_location = f"{chr(65 + col_idx - 1)}{row_idx}"
                if i:
                    ws.cell(row=result_row+4, column=j, value=value)
                elif not insert_image(value, cell_location):
                    ws.cell(row=result_row+3, column=j, value=value)

    for i, data in enumerate(result_data[0]):
        row_idx, col_idx = result_data[1][i]
        cell_location = f"{chr(65 + col_idx - 1)}{row_idx}"
        if not insert_image(data, cell_location):
            ws.cell(row=row_idx, column=col_idx, value=data)
            

q = deployment_input("상대 턴 움직임을 만드시겠습니까? (넘어가려면 빈 입력): ", 1, my_deployment)
if q:
    ws.cell(row=result_row+5, column=1, value="상대")
    insert_deployment(result_row+5, 2)

# 열의 너비를 15로 설정
for col_letter in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
    ws.column_dimensions[col_letter].width = 15

# 카드 이미지 들어갈 위치만 130사이즈로 조절
for row in range(result_row, result_row + 4):
    ws.row_dimensions[row].height = 130
ws.row_dimensions[2].height = 130
for enu, row in enumerate(ws.iter_rows(min_row=result_row, max_row=result_row + 5, min_col=1, max_col=ws.max_column), start=1):
    for cell in row:
        cell.alignment = alignment_style
        if enu == 5:
            cell.font = font_style_text
        else:
            cell.font = font_style_card
for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = alignment_style
        cell.font = font_style_card

# 1행 수정
ws.merge_cells('A1:O1')
ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

wb.save(file_path)

my_deployment.append('')
deployment_save("Sheet", my_deployment)
print("저장 완료")
